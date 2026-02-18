# -*- coding: utf-8 -*-
"""
Created on Mon Sep  1 22:10:06 2025

@author: nguye
"""
import numpy as np
from numba import njit
from numpy.polynomial.legendre import leggauss
# from shape_function_derivative import dNi_dxi, dNi_deta


# ================== SHAPE FUNCTIONS ==================
@njit(fastmath=True, cache=True)
def gamma(eta, m, n):
    return np.abs(1 - (1 - ((eta + 1) / 2)**2)**m)**n

@njit(fastmath=True, cache=True)
def NAG3_one(eta, m, n):
    G0 = gamma(0, m, n)
    G_eta = gamma(eta, m, n)
    return (-G0 + (G0 - 1)*eta + G_eta) / (1 - 2 * G0)

@njit(fastmath=True, cache=True)
def NAG3_two(eta, m, n):
    G0 = gamma(0, m, n)
    G_eta = gamma(eta, m, n)
    return (1 + eta - 2 * G_eta) / (1 - 2 * G0)

@njit(fastmath=True, cache=True)
def NAG3_three(eta, m, n):
    G0 = gamma(0, m, n)
    G_eta = gamma(eta, m, n)
    return (-G0 - G0 * eta + G_eta) / (1 - 2 * G0)

@njit(fastmath=True, cache=True)
def shape_functions_AGc3(xi, eta, m1, n1, m2, n2):
    N = np.zeros(9)
    N[0] = NAG3_one(xi, m1, n1)   * NAG3_one(eta, m2, n2)
    N[1] = NAG3_three(xi, m1, n1) * NAG3_one(eta, m2, n2)
    N[2] = NAG3_three(xi, m1, n1) * NAG3_three(eta, m2, n2)
    N[3] = NAG3_one(xi, m1, n1)   * NAG3_three(eta, m2, n2)
    N[4] = NAG3_two(xi, m1, n1)   * NAG3_one(eta, m2, n2)
    N[5] = NAG3_three(xi, m1, n1) * NAG3_two(eta, m2, n2)
    N[6] = NAG3_two(xi, m1, n1)   * NAG3_three(eta, m2, n2)
    N[7] = NAG3_one(xi, m1, n1)   * NAG3_two(eta, m2, n2)
    N[8] = NAG3_two(xi, m1, n1)   * NAG3_two(eta, m2, n2)
    return N

@njit(fastmath=True, cache=True)
def shape_functions_AGc4(xi, eta, m1, n1, m2, n2):
    N = np.zeros(9)
    N[0] = NAG3_three(-xi, m1, n1)   * NAG3_one(eta, m2, n2)
    N[1] = NAG3_one(-xi, m1, n1) * NAG3_one(eta, m2, n2)
    N[2] = NAG3_one(-xi, m1, n1) * NAG3_three(eta, m2, n2)
    N[3] = NAG3_three(-xi, m1, n1)   * NAG3_three(eta, m2, n2)
    N[4] = NAG3_two(-xi, m1, n1)   * NAG3_one(eta, m2, n2)
    N[5] = NAG3_one(-xi, m1, n1) * NAG3_two(eta, m2, n2)
    N[6] = NAG3_two(-xi, m1, n1)   * NAG3_three(eta, m2, n2)
    N[7] = NAG3_three(-xi, m1, n1)   * NAG3_two(eta, m2, n2)
    N[8] = NAG3_two(-xi, m1, n1)   * NAG3_two(eta, m2, n2)
    return N

@njit(fastmath=True, cache=True)
def dN_dxi_AGc3(xi, eta, m1, n1, m2, n2):
    G01 = np.abs(1 - (3/4)**m1)**n1
    G02 = np.abs(1 - (3/4)**m2)**n2
    denom1 = 1 - 2*G01
    denom2 = 1 - 2*G02
    # denom1_sq = denom1**2

    # Helper for exponentiation, matches lambda code behavior:
    def pow_safe(base, exp):
        if exp == 0:
            return 1.0
        else:
            return base**exp

    # Lambda code mimics:
    # ...* (1 - 1/4*(1 + xi)**2)**(m-1) * np.abs(...)**(n-1) * np.sign(...)
    # For m=1, the (m-1) term is zero, so just 1.0
    # For n=1, (n-1) = 0, so just 1.0

    term_xi = 1 - 1/4*(1 + xi)**2
    term_eta = 1 - 1/4*(1 + eta)**2
    inner_xi = 1 - term_xi**m1
    # abs_inner_xi = np.abs(inner_xi)
    sign_inner_xi = np.sign(inner_xi)

    def aux_xi(xi_):
        t = 1 - 1/4*(1 + xi_)**2
        val = 1 - t**m1
        # (m1-1) and (n1-1)
        base1 = pow_safe(t, m1-1)
        base2 = pow_safe(np.abs(val), n1-1)
        sign = np.sign(val)
        return base1 * base2 * sign

    # For each derivative, we’ll apply this logic:
    # [ref] (-1 + G01 - 0.5*m1*n1*(-1-xi)*base1*base2*sign)
    # Note: The aux_xi(xi) here is base1*base2*sign
    base1 = pow_safe(term_xi, m1-1)
    base2 = pow_safe(np.abs(inner_xi), n1-1)
    aux_term = base1 * base2 * sign_inner_xi

    results = np.zeros(9)

    # dN1/dxi
    results[0] =  (-G02 + eta * (-1 + G02) + np.abs(1 - term_eta**m2)**n2
                   )  * (-1 + G01 - 0.5*m1*n1*(-1 - xi)*aux_term) / (denom1 * denom2)

    # dN2/dxi
    results[1] = (
        (-G02 + eta*(-1 + G02) + np.abs(1 - term_eta**m2)**n2)
        * (-G01 - 0.5*m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN3/dxi
    results[2] = (
        (-G02 - eta*G02 + np.abs(1 - term_eta**m2)**n2)
        * (-G01 - 0.5*m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN4/dxi
    results[3] = (
        (-G02 - eta*G02 + np.abs(1 - term_eta**m2)**n2)
        * (-1 + G01 - 0.5*m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN5/dxi
    results[4] = (
        (-G02 + eta*(-1 + G02) + np.abs(1 - term_eta**m2)**n2)
        * (1 + m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN6/dxi
    results[5] = (
        (1 + eta - 2*np.abs(1 - term_eta**m2)**n2)
        * (-G01 - 0.5*m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN7/dxi
    results[6] = (
        (-G02 - eta*G02 + np.abs(1 - term_eta**m2)**n2)
        * (1 + m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN8/dxi
    results[7] = (
        (1 + eta - 2*np.abs(1 - term_eta**m2)**n2)
        * (-1 + G01 - 0.5*m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    # dN9/dxi
    results[8] = (
        (1 + eta - 2*np.abs(1 - term_eta**m2)**n2)
        * (1 + m1*n1*(-1 - xi)*aux_term)
        / (denom1 * denom2)
    )

    return results


@njit(fastmath=True, cache=True)
def dN_deta_AGc3(xi, eta, m1, n1, m2, n2):
    G01 = np.abs(1 - (3/4)**m1)**n1
    G02 = np.abs(1 - (3/4)**m2)**n2
    denom1 = 1 - 2*G01
    denom2 = 1 - 2*G02
    # denom1_sq = denom1**2

    # Safe exponentiation for special cases
    def pow_safe(base, exp):
        if exp == 0:
            return 1.0
        else:
            return base**exp

    term_xi = 1 - 1/4*(1 + xi)**2
    term_eta = 1 - 1/4*(1 + eta)**2
    inner_eta = 1 - term_eta**m2
    # abs_inner_eta = np.abs(inner_eta)
    sign_inner_eta = np.sign(inner_eta)

    # Helper for base^exp * |inner|^exp * sign(inner)
    def aux_eta(eta_):
        t = 1 - 1/4*(1 + eta_)**2
        val = 1 - t**m2
        base1 = pow_safe(t, m2-1)
        base2 = pow_safe(np.abs(val), n2-1)
        sign = np.sign(val)
        return base1 * base2 * sign

    base1 = pow_safe(term_eta, m2-1)
    base2 = pow_safe(np.abs(inner_eta), n2-1)
    aux_term = base1 * base2 * sign_inner_eta

    results = np.zeros(9)

    # dN1/deta
    results[0] =(-G01 + xi*(-1 + G01) + np.abs(1 - term_xi**m1)**n1
    ) * (-1 + G02 - 0.5*m2*n2*(-1 - eta)*aux_term)/(denom1 * denom2)

    # dN2/deta
    results[1] = (
        (-G01 - xi*G01 + np.abs(1 - term_xi**m1)**n1)
        * (-1 + G02 - 0.5*m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN3/deta
    results[2] = (
        (-G01 - xi*G01 + np.abs(1 - term_xi**m1)**n1)
        * (-G02 - 0.5*m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN4/deta
    results[3] = (
        (-G01 + xi*(-1 + G01) + np.abs(1 - term_xi**m1)**n1)
        * (-G02 - 0.5*m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN5/deta
    results[4] = (
        (1 + xi - 2*np.abs(1 - term_xi**m1)**n1)
        * (-1 + G02 - 0.5*m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN6/deta
    results[5] = (
        (-G01 - xi*G01 + np.abs(1 - term_xi**m1)**n1)
        * (1 + m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN7/deta
    results[6] = (
        (1 + xi - 2*np.abs(1 - term_xi**m1)**n1)
        * (-G02 - 0.5*m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN8/deta
    results[7] = (
        (-G01 + xi*(-1 + G01) + np.abs(1 - term_xi**m1)**n1)
        * (1 + m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    # dN9/deta
    results[8] = (
        (1 + xi - 2*np.abs(1 - term_xi**m1)**n1)
        * (1 + m2*n2*(-1 - eta)*aux_term)
        / (denom1 * denom2)
    )

    return results

@njit(fastmath=True)
def _sgn(x):
    if x > 0.0:
        return 1.0
    elif x < 0.0:
        return -1.0
    else:
        return 0.0

@njit(fastmath=True)
def dN_dxi_AGc4(eta1, eta2, m1, n1, m2, n2):
    # local sign (Numba-friendly)
    def _sgn(x):
        return 1.0 if x > 0.0 else (-1.0 if x < 0.0 else 0.0)

    # A1 = |1 - (3/4)^m1|^n1,  A2 = |1 - (3/4)^m2|^n2
    A1 = abs(1.0 - (0.75)**m1)**n1
    A2 = abs(1.0 - (0.75)**m2)**n2

    # Gp2(eta2) = |1 - (1 - 1/4 (1 + eta2)^2)^m2|^n2  (η2-side family m2)
    base_eta2 = 1.0 - 0.25*(1.0 + eta2)**2
    gp2_inner = 1.0 - base_eta2**m2
    Gp2 = abs(gp2_inner)**n2

    # η2-side linear combos (LEFT factor in your Mathematica rows; all m2)
    # (-A2 + η2(-1 + A2) + Gp2)
    S2a = -A2 + eta2*( -1.0 + A2 ) + Gp2
    # (-A2 - η2*A2 + Gp2)
    S2b = -A2 - eta2*A2 + Gp2
    # (1 + η2 - 2*Gp2)
    S2c =  1.0 + eta2 - 2.0*Gp2

    # η1-side derivative pieces (RIGHT factor in the Mathematica rows; family m1)
    base_eta1 = 1.0 - 0.25*(1.0 - eta1)**2              # (1 - 1/4 (1 - η1)^2)
    h1_inner  = 1.0 - base_eta1**m1                      # 1 - base^m1
    base_pow  = base_eta1**(m1 - 1.0)                    # base^(m1-1)
    abs_pow   = abs(h1_inner)**(n1 - 1.0)                # |h1|^(n1-1)
    sig       = _sgn(h1_inner)
    # common derivative prefactors:  (1 - η1) appears exactly as in the formula
    T1_half = 0.5 * m1 * n1 * base_pow * (1.0 - eta1) * abs_pow * sig
    T1_full =       m1 * n1 * base_pow * (1.0 - eta1) * abs_pow * sig

    # η1-side combinations (three variants)
    # (A1 - 1/2 * ...)
    fA = (A1 - T1_half)
    # (1 - A1 - 1/2 * ...)
    gA = (1.0 - A1 - T1_half)
    # (-1 + ...)
    hA = (-1.0 + T1_full)

    # denominators
    denom1 = 1.0 - 2.0*A1
    denom2 = 1.0 - 2.0*A2
    invden = 1.0 / (denom1 * denom2)

    # Map to the 3x3 grid (row-major) exactly as in the Mathematica output:
    # Row 1:  (S2a * fA), (S2a * gA), (S2b * gA)
    e1 = (S2a * fA) * invden
    e2 = (S2a * gA) * invden
    e3 = (S2b * gA) * invden
    # Row 2:  (S2b * fA), (S2a * hA), (S2c * gA)
    e4 = (S2b * fA) * invden
    e5 = (S2a * hA) * invden
    e6 = (S2c * gA) * invden
    # Row 3:  (S2b * hA), (S2c * fA), (S2c * hA)
    e7 = (S2b * hA) * invden
    e8 = (S2c * fA) * invden
    e9 = (S2c * hA) * invden

    out = np.empty(9, dtype=np.float64)
    out[0] = e1; out[1] = e2; out[2] = e3
    out[3] = e4; out[4] = e5; out[5] = e6
    out[6] = e7; out[7] = e8; out[8] = e9
    return out


# @njit(fastmath=True)
# def dN_dxi_AGc4(eta1, eta2, m1, n1, m2, n2):
#     # |1 - (3/4)^m|^n terms (G0^n)
#     g0_1_inner = 1.0 - (0.75)**m1
#     g0_2_inner = 1.0 - (0.75)**m2
#     A1 = abs(g0_1_inner)**n1  # Abs[1 - (3/4)^m1]^n1
#     A2 = abs(g0_2_inner)**n2  # Abs[1 - (3/4)^m2]^n2

#     # |1 - (1 - 1/4 (1 + eta2)^2)^m|^n terms (at +eta2)
#     gp1_inner = 1.0 - (1.0 - 0.25*(1.0 + eta2)**2)**m1
#     gp2_inner = 1.0 - (1.0 - 0.25*(1.0 + eta2)**2)**m2
#     Gp1 = abs(gp1_inner)**n1
#     Gp2 = abs(gp2_inner)**n2

#     # Pieces depending on eta1 for the derivative factors
#     base1 = 1.0 - 0.25*(1.0 - eta1)**2              # (1 - 1/4 (1 - eta1)^2)
#     h1_inner = 1.0 - base1**m1                       # 1 - (1 - 1/4 (1 - eta1)^2)^m1
#     base1_pow = base1**(m1 - 1.0)                    # (1 - 1/4 (1 - eta1)^2)^(m1 - 1)
#     habs_pow = abs(h1_inner)**(n1 - 1.0)             # Abs[...]^(n1 - 1)
#     sig = _sgn(h1_inner)                              # Sign[...]

#     # Common combos from your formula
#     T1_half = 0.5 * m1 * n1 * base1_pow * (1.0 - eta1) * habs_pow * sig
#     T1_full =       m1 * n1 * base1_pow * (1.0 - eta1) * habs_pow * sig

#     L1  = -A1 + eta2*(-1.0 + A1) + Gp1                 # (-A1 + eta2(-1 + A1) + Gp1)
#     S2a = -A2 + eta2*(-1.0 + A2) + Gp2                 # (-A2 + eta2(-1 + A2) + Gp2)
#     S2b = -A2 - eta2*A2 + Gp2                          # (-A2 - eta2*A2 + Gp2)
#     S2c =  1.0 + eta2 - 2.0*Gp2                        # (1 + eta2 - 2*Gp2)

#     fA = (A1 - T1_half)                                # (A1 - 1/2 * ...)
#     gA = (1.0 - A1 - T1_half)                          # (1 - A1 - 1/2 * ...)
#     hA = (-1.0 + T1_full)                              # (-1 + ...)

#     denom1 = 1.0 - 2.0*A1
#     denom2 = 1.0 - 2.0*A2

#     # The nine expressions, in order
#     e1 = (L1  * fA) / (denom1 * denom2)
#     e2 = (S2a * gA) / (denom1 * denom2)
#     e3 = (S2b * gA) / (denom1 * denom2)
#     e4 = (S2b * fA) / (denom1 * denom2)
#     e5 = (S2a * hA) / (denom1 * denom2)
#     e6 = (S2c * gA) / (denom1 * denom2)
#     e7 = (S2b * hA) / (denom1 * denom2)
#     e8 = (S2c * fA) / (denom1 * denom2)
#     e9 = (S2c * hA) / (denom1 * denom2)

#     out = np.empty(9, dtype=np.float64)
#     out[0] = e1; out[1] = e2; out[2] = e3
#     out[3] = e4; out[4] = e5; out[5] = e6
#     out[6] = e7; out[7] = e8; out[8] = e9
#     return out


@njit(fastmath=True)
def dN_deta_AGc4(eta1, eta2, m1, n1, m2, n2):
    def _sgn(x):
        return 1.0 if x > 0.0 else (-1.0 if x < 0.0 else 0.0)

    # A1 = |1 - (3/4)^m1|^n1,  A2 = |1 - (3/4)^m2|^n2
    A1 = abs(1.0 - (0.75)**m1)**n1
    A2 = abs(1.0 - (0.75)**m2)**n2

    # Gm1(eta1) = |1 - (1 - 1/4 (1 - eta1)^2)^m1|^n1
    base1_m1 = 1.0 - 0.25*(1.0 - eta1)**2
    h1_m1_inner = 1.0 - base1_m1**m1
    Gm1 = abs(h1_m1_inner)**n1

    # η2-side (always m2-family in your closed form)
    base2 = 1.0 - 0.25*(1.0 + eta2)**2
    h2_m2_inner = 1.0 - base2**m2
    fac_m2_half = 0.5 * m2 * n2 * (-1.0 - eta2) * (base2**(m2 - 1.0)) \
                  * (abs(h2_m2_inner)**(n2 - 1.0)) * _sgn(h2_m2_inner)
    fac_m2_full = m2 * n2 * (-1.0 - eta2) * (base2**(m2 - 1.0)) \
                  * (abs(h2_m2_inner)**(n2 - 1.0)) * _sgn(h2_m2_inner)

    # Left (η1) combos
    L_a = -A1 + eta1*A1 + Gm1                          # (-A1 + η1 A1 + Gm1)
    L_b = -A1 + eta1*(1.0 - A1) + Gm1                  # (-A1 - η1(-1 + A1) + Gm1)
    L_c = 1.0 - eta1 - 2.0*Gm1                         # (1 - η1 - 2 Gm1)

    # Right (η2) combos — all with A2,m2
    R1   = (-1.0 + A2 - fac_m2_half)                   # matches first row
    R2_a = (-1.0 + A2 - fac_m2_half)
    R2_b = (-A2 - fac_m2_half)
    R2_c = (1.0 + fac_m2_full)

    denom1 = 1.0 - 2.0*A1
    denom2 = 1.0 - 2.0*A2
    invden = 1.0 / (denom1 * denom2)

    e1 = (L_a * R1)   * invden
    e2 = (L_b * R2_a) * invden
    e3 = (L_b * R2_b) * invden
    e4 = (L_a * R2_b) * invden
    e5 = (L_c * R2_a) * invden
    e6 = (L_b * R2_c) * invden
    e7 = (L_c * R2_b) * invden
    e8 = (L_a * R2_c) * invden
    e9 = (L_c * R2_c) * invden

    out = np.empty(9, dtype=np.float64)
    out[0] = e1; out[1] = e2; out[2] = e3
    out[3] = e4; out[4] = e5; out[5] = e6
    out[6] = e7; out[7] = e8; out[8] = e9
    return out

# @njit(fastmath=True)
# def dN_deta_AGc4(eta1, eta2, m1, n1, m2, n2):
#     # A1 = Abs[1 - (3/4)^m1]^n1, A2 = Abs[1 - (3/4)^m2]^n2
#     g0_1_inner = 1.0 - (0.75)**m1
#     g0_2_inner = 1.0 - (0.75)**m2
#     A1 = abs(g0_1_inner)**n1
#     A2 = abs(g0_2_inner)**n2

#     # Gm1 = Abs[1 - (1 - 1/4 (1 - eta1)^2)^m1]^n1   (depends on eta1)
#     base1_m1 = 1.0 - 0.25*(1.0 - eta1)**2
#     h1_m1_inner = 1.0 - base1_m1**m1
#     Gm1 = abs(h1_m1_inner)**n1

#     # Eta2-derivative pieces (1 + eta2) side, for m1 and m2 families
#     base2_m1 = 1.0 - 0.25*(1.0 + eta2)**2
#     base2_m2 = 1.0 - 0.25*(1.0 + eta2)**2  # same functional form with m2
#     h2_m1_inner = 1.0 - base2_m1**m2
#     h2_m2_inner = 1.0 - base2_m2**m2

#     # Common factors: 0.5 * m*n * (-1 - eta2) * base^(m-1) * Abs(h)^(n-1) * Sign(h)
#     # (these appear as "- 1/2 m n (-1 - eta2) ..." in your formula)
#     fac_m1_half = 0.5 * m2 * n2 * (-1.0 - eta2) \
#                   * (base2_m1**(m2 - 1.0)) \
#                   * (abs(h2_m1_inner)**(n2 - 1.0)) \
#                   * _sgn(h2_m1_inner)

#     fac_m2_half = 0.5 * m2 * n2 * (-1.0 - eta2) \
#                   * (base2_m2**(m2 - 1.0)) \
#                   * (abs(h2_m2_inner)**(n2 - 1.0)) \
#                   * _sgn(h2_m2_inner)

#     # Also appears as "+ m2 n2 (-1 - eta2) ..." (without 1/2)
#     fac_m2_full = m2 * n2 * (-1.0 - eta2) \
#                   * (base2_m2**(m2 - 1.0)) \
#                   * (abs(h2_m2_inner)**(n2 - 1.0)) \
#                   * _sgn(h2_m2_inner)

#     # Linear-in-eta1 combos (your first parentheses group variants)
#     # (-A1 + eta1*A1 + Gm1)
#     L_a = -A1 + eta1*A1 + Gm1
#     # (-A1 - eta1*(-1 + A1) + Gm1) == -A1 + eta1*(1 - A1) + Gm1
#     L_b = -A1 + eta1*(1.0 - A1) + Gm1
#     # (1 - eta1 - 2*Gm1)
#     L_c = 1.0 - eta1 - 2.0*Gm1

#     # Eta2-side combinations (your second parentheses group variants)
#     # With m1 family (only in the very first term):
#     R1 = (-1.0 + A1 - fac_m1_half)

#     # With m2 family (used in most other terms):
#     R2_a = (-1.0 + A2 - fac_m2_half)
#     R2_b = (-A2 - fac_m2_half)
#     R2_c = (1.0 + fac_m2_full)

#     # Denominators
#     denom1 = 1.0 - 2.0*A1
#     denom2 = 1.0 - 2.0*A2

#     # Nine expressions, in the same order as provided
#     e1 = (L_a * R1)         / (denom1 * denom2)
#     e2 = (L_b * R2_a)       / (denom1 * denom2)
#     e3 = (L_b * R2_b)       / (denom1 * denom2)
#     e4 = (L_a * R2_b)       / (denom1 * denom2)
#     e5 = (L_c * R2_a)       / (denom1 * denom2)
#     e6 = (L_b * R2_c)       / (denom1 * denom2)
#     e7 = (L_c * R2_b)       / (denom1 * denom2)
#     e8 = (L_a * R2_c)       / (denom1 * denom2)
#     e9 = (L_c * R2_c)       / (denom1 * denom2)

#     out = np.empty(9, dtype=np.float64)
#     out[0] = e1; out[1] = e2; out[2] = e3
#     out[3] = e4; out[4] = e5; out[5] = e6
#     out[6] = e7; out[7] = e8; out[8] = e9
#     return out
# ================== Q9 GEOMETRY (bi-quadratic Lagrange) ==================
@njit(fastmath=True, cache=True)
def q9_1d_L(ξ):
    # Nodes at ξ = -1, 0, +1
    L1 = 0.5 * ξ * (ξ - 1.0)   # N at -1
    L2 = 1.0 - ξ*ξ             # N at  0
    L3 = 0.5 * ξ * (ξ + 1.0)   # N at +1
    return L1, L2, L3

@njit(fastmath=True, cache=True)
def q9_1d_dL(ξ):
    dL1 = ξ - 0.5
    dL2 = -2.0 * ξ
    dL3 = ξ + 0.5
    return dL1, dL2, dL3

@njit(fastmath=True, cache=True)
def q9_geom_shapes_derivs(xi, eta):
    """
    Q9 geometry shapes N^g_a, and derivatives wrt (ξ,η).
    Ordering (a=0..8):
      0:(-1,-1)=L1(ξ)L1(η)
      1:(+1,-1)=L3(ξ)L1(η)
      2:(+1,+1)=L3(ξ)L3(η)
      3:(-1,+1)=L1(ξ)L3(η)
      4:( 0,-1)=L2(ξ)L1(η)
      5:(+1, 0)=L3(ξ)L2(η)
      6:( 0,+1)=L2(ξ)L3(η)
      7:(-1, 0)=L1(ξ)L2(η)
      8:( 0, 0)=L2(ξ)L2(η)
    """
    Lx1, Lx2, Lx3 = q9_1d_L(xi)
    Ly1, Ly2, Ly3 = q9_1d_L(eta)
    dLx1, dLx2, dLx3 = q9_1d_dL(xi)
    dLy1, dLy2, dLy3 = q9_1d_dL(eta)

    N = np.empty(9)
    dN_dxi  = np.empty(9)
    dN_deta = np.empty(9)

    # corners
    N[0] = Lx1*Ly1; dN_dxi[0] = dLx1*Ly1; dN_deta[0] = Lx1*dLy1
    N[1] = Lx3*Ly1; dN_dxi[1] = dLx3*Ly1; dN_deta[1] = Lx3*dLy1
    N[2] = Lx3*Ly3; dN_dxi[2] = dLx3*Ly3; dN_deta[2] = Lx3*dLy3
    N[3] = Lx1*Ly3; dN_dxi[3] = dLx1*Ly3; dN_deta[3] = Lx1*dLy3
    # midsides
    N[4] = Lx2*Ly1; dN_dxi[4] = dLx2*Ly1; dN_deta[4] = Lx2*dLy1
    N[5] = Lx3*Ly2; dN_dxi[5] = dLx3*Ly2; dN_deta[5] = Lx3*dLy2
    N[6] = Lx2*Ly3; dN_dxi[6] = dLx2*Ly3; dN_deta[6] = Lx2*dLy3
    N[7] = Lx1*Ly2; dN_dxi[7] = dLx1*Ly2; dN_deta[7] = Lx1*dLy2
    # center
    N[8] = Lx2*Ly2; dN_dxi[8] = dLx2*Ly2; dN_deta[8] = Lx2*dLy2

    return N, dN_dxi, dN_deta

# ================== CONSTITUTIVE (Voigt 3x3) ==================
@njit(fastmath=True, cache=True)
def constitutive_matrix(mu, nu, planestrain):
    E = 2.0 * mu * (1.0 + nu)
    D = np.zeros((3, 3))
    if planestrain == 1:
        lam = 2.0 * mu * nu / (1.0 - 2.0 * nu)
        D[0,0] = lam + 2.0*mu; D[0,1] = lam
        D[1,0] = lam;           D[1,1] = lam + 2.0*mu
        D[2,2] = mu
    else:
        c = E / (1.0 - nu*nu)
        D[0,0] = c;    D[0,1] = c*nu
        D[1,0] = c*nu; D[1,1] = c
        D[2,2] = c*(1.0 - nu)*0.5
    return D

# ================== B-matrix (2D, eng shear) ==================
@njit(fastmath=True, cache=True)
def build_B(dNdx, dNdy):
    n = dNdx.shape[0]
    B = np.zeros((3, 2*n))
    for a in range(n):
        ia = 2*a
        B[0, ia]     = dNdx[a]
        B[1, ia + 1] = dNdy[a]
        B[2, ia]     = dNdy[a]
        B[2, ia + 1] = dNdx[a]
    return B

# ================== ELEMENT STIFFNESS (AG field, Q9 geometry) ==================
@njit(fastmath=True, cache=True)
def ele_stiffness_numba_Q9geo_AGfield(
    coords,                # (9,2) geometry nodes in Q9 order (see above)
    m1, n1, m2, n2, AGtype,        # AG field parameters
    xi_pts, xi_wts,        # Gauss ξ
    eta_pts, eta_wts,      # Gauss η
    ndof,                  # 2
    mu, nu, planestrain    # material
):
    n_nodes = coords.shape[0]  # expect 9
    Ke = np.zeros((ndof * n_nodes, ndof * n_nodes))
    D = constitutive_matrix(mu, nu, planestrain)
    for i in range(xi_pts.shape[0]):
        xi = xi_pts[i]; wxi = xi_wts[i]
        for j in range(eta_pts.shape[0]):
            eta = eta_pts[j]; weta = eta_wts[j]
            
            if AGtype == 2:      
                # --- Field derivatives wrt parent: AG/Q9 you defined ---
                dN_dxi  = dN_dxi_AGc4(xi, eta, m1, n1, m2, n2)   # (9,)
                dN_deta = dN_deta_AGc4(xi, eta, m1, n1, m2, n2)  # (9,)
            # elif AGtype == 1: 
            #     # --- Field derivatives wrt parent: AG/Q9 you defined ---
            #     dN_dxi  = dN_dxi_AGc3(xi, eta, m1, n1, m2, n2)   # (9,)
            #     dN_deta = dN_deta_AGc3(xi, eta, m1, n1, m2, n2)  # (9,)
            else:
                # --- Field derivatives wrt parent: AG/Q9 you defined ---
                dN_dxi  = dN_dxi_AGc3(xi, eta, m1, n1, m2, n2)   # (9,)
                dN_deta = dN_deta_AGc3(xi, eta, m1, n1, m2, n2)  # (9,)
                
            # --- Geometry from standard Q9 shapes (IMPORTANT) ---
            Ng, dNg_dxi, dNg_deta = q9_geom_shapes_derivs(xi, eta)  # (9,)
            # Build Jacobian from geometry derivatives
            dx_dxi  = 0.0; dx_deta = 0.0; dy_dxi = 0.0; dy_deta = 0.0
            for a in range(9):
                x_a = coords[a, 0]; y_a = coords[a, 1]
                dx_dxi  += dNg_dxi[a]  * x_a
                dx_deta += dNg_deta[a] * x_a
                dy_dxi  += dNg_dxi[a]  * y_a
                dy_deta += dNg_deta[a] * y_a

            J00 = dx_dxi; J01 = dx_deta
            J10 = dy_dxi; J11 = dy_deta
            detJ = J00*J11 - J01*J10
            invDet = 1.0 / detJ
            Jinv00 =  J11*invDet; Jinv01 = -J01*invDet
            Jinv10 = -J10*invDet; Jinv11 =  J00*invDet

            # --- Map AG field derivatives to x,y ---
            dNdx = np.empty(9)
            dNdy = np.empty(9)
            for a in range(9):
                dNdx[a] = Jinv00 * dN_dxi[a] + Jinv01 * dN_deta[a]
                dNdy[a] = Jinv10 * dN_dxi[a] + Jinv11 * dN_deta[a]

            # --- B and stiffness contribution ---
            B  = build_B(dNdx, dNdy)
            Bt = B.T
            w  = wxi * weta
            Ke += (Bt @ D @ B) * detJ * w

    return Ke

def ele_stiffness_Q9geo_AGfield(coords, eletype, xi_pts, xi_wts, eta_pts, eta_wts, ndof, materialprops):
    m1, n1, m2, n2, ngauss, AGtype = eletype
    mu, nu, planestrain = materialprops
    return ele_stiffness_numba_Q9geo_AGfield(coords, m1, n1, m2, n2, AGtype, xi_pts, xi_wts, eta_pts, eta_wts, ndof, mu, nu, planestrain)



#### export to excel routine
from pathlib import Path
from openpyxl import Workbook
def _script_dir():
    """
    Directory of this .py file. If __file__ is not defined (notebook/interactive),
    use current working directory.
    """
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()

def _write_Ke_sheet(wb, e, conn, eletype, Ke):
    """
    Write element stiffness Ke (18x18 for Q9 2D) to a new worksheet.
    """
    # Excel sheet name max length is 31
    sheet_name = f"E{e+1:05d}"  # 1-based element indexing in name
    ws = wb.create_sheet(title=sheet_name)

    # Header info
    ws.append(["Element ID (0-based)", int(e)])
    ws.append(["Element ID (1-based)", int(e + 1)])
    ws.append(["Connectivity (global node ids, Q9 order)"] + [int(x) for x in conn])
    ws.append(["etype = [m1,n1,m2,n2,ngauss,AGtype]"] + [float(x) for x in eletype])
    ws.append([])

    # DOF labels: n1_ux, n1_uy, ..., n9_ux, n9_uy
    dof_labels = []
    for a in range(9):
        dof_labels.append(f"n{a+1}_ux")
        dof_labels.append(f"n{a+1}_uy")

    # Table header row
    ws.append([""] + dof_labels)

    # Data rows
    # Ke is (18,18)
    for i in range(Ke.shape[0]):
        ws.append([dof_labels[i]] + [float(v) for v in Ke[i, :]])

    # Optional: freeze pane at first data cell
    ws.freeze_panes = "B7"


# ================== GLOBAL STIFFNESS (dense for clarity) ==================
# def global_stiffness_Q9geo_AGfield(
#     coords, connectivity,     # coords:(N,2); connectivity:(Ne,9) in Q9 order
#     etype,     # AG params [m1,n1,m2,n2,ngauss] for selected elems
#     ndof, materialprops
# ):
#     Nnodes = coords.shape[0]
#     Kglob  = np.zeros((Nnodes*ndof, Nnodes*ndof))
#     Ne = connectivity.shape[0]

#     for e in range(Ne):
#         conn   = connectivity[e]
#         ecoord = coords[conn, :]  # (9,2)

#         ngauss  = etype[e][4]
#         xi_pts, xi_wts   = leggauss(ngauss)
#         eta_pts, eta_wts = leggauss(ngauss)

#         Ke = ele_stiffness_Q9geo_AGfield(ecoord, etype[e] , xi_pts, xi_wts, eta_pts, eta_wts, ndof, materialprops)
        
        
#         # assemble into Kglob
#         edofs = np.empty(2*conn.shape[0], dtype=np.int64)
#         p = 0
#         for a in conn:
#             edofs[p]   = 2*a
#             edofs[p+1] = 2*a + 1
#             p += 2

#         for ii in range(edofs.shape[0]):
#             I = edofs[ii]
#             for jj in range(edofs.shape[0]):
#                 J = edofs[jj]
#                 Kglob[I, J] += Ke[ii, jj]

#     return Kglob


def global_stiffness_Q9geo_AGfield(
    coords, connectivity,     # coords:(N,2); connectivity:(Ne,9) in Q9 order
    etype,     # AG params [m1,n1,m2,n2,ngauss,AGtype]
    ndof, materialprops
):
    Nnodes = coords.shape[0]
    Kglob  = np.zeros((Nnodes*ndof, Nnodes*ndof))
    Ne = connectivity.shape[0]

    # ============================================================
    # OPTIONAL EXPORT (comment this whole block out when not needed)
    # ============================================================
    try:
        script_dir = Path(__file__).resolve().parent
    except NameError:
        script_dir = Path.cwd()

    out_dir = script_dir / "element_stiffness"
    out_dir.mkdir(parents=True, exist_ok=True)

    export_path = out_dir / "Ke_by_element.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    # ============================================================

    for e in range(Ne):
        conn   = connectivity[e]
        ecoord = coords[conn, :]  # (9,2)

        ngauss  = int(etype[e][4])
        xi_pts, xi_wts   = leggauss(ngauss)
        eta_pts, eta_wts = leggauss(ngauss)

        Ke = ele_stiffness_Q9geo_AGfield(
            ecoord, etype[e], xi_pts, xi_wts, eta_pts, eta_wts, ndof, materialprops
        )

        # ============================================================
        # OPTIONAL EXPORT PER ELEMENT (comment out when not needed)
        # ============================================================
        sheet_name = f"E{e+1:05d}"
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Element ID (0-based)", int(e)])
        ws.append(["Element ID (1-based)", int(e + 1)])
        ws.append(["Connectivity (global node ids, Q9 order)"] + [int(x) for x in conn])
        ws.append(["etype = [m1,n1,m2,n2,ngauss,AGtype]"] + [float(x) for x in etype[e]])
        ws.append([])

        dof_labels = []
        for a in range(9):
            dof_labels.append(f"n{a+1}_ux")
            dof_labels.append(f"n{a+1}_uy")

        ws.append([""] + dof_labels)
        for i in range(Ke.shape[0]):
            ws.append([dof_labels[i]] + [float(v) for v in Ke[i, :]])

        ws.freeze_panes = "B7"
        # ============================================================

        # assemble into Kglob
        edofs = np.empty(2*conn.shape[0], dtype=np.int64)
        p = 0
        for a in conn:
            edofs[p]   = 2*a
            edofs[p+1] = 2*a + 1
            p += 2

        for ii in range(edofs.shape[0]):
            I = edofs[ii]
            for jj in range(edofs.shape[0]):
                J = edofs[jj]
                Kglob[I, J] += Ke[ii, jj]

    # ============================================================
    # OPTIONAL SAVE (comment out when not needed)
    # ============================================================
    wb.save(export_path)
    print(f"[OK] Saved element stiffness matrices to: {export_path}")
    # ============================================================

    return Kglob


#%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%


# ---- small helper: element dof indices [ux1,uy1,...,ux9,uy9]
def _element_edofs(conn):
    edofs = np.empty(2*len(conn), dtype=np.int64)
    p = 0
    for n in conn:
        edofs[p]   = 2*n
        edofs[p+1] = 2*n + 1
        p += 2
    return edofs

def element_gauss_strain_stress_with_deformed(coords, conn, u, eletype, materialprops):
    """
    ONE Q9 element: compute strain, stress, sigma_zz, von Mises, and both initial & deformed GP coords.

    Returns dict:
      'xi_eta'   : (ng^2, 2)
      'xy0'      : (ng^2, 2)   initial GP coords
      'xy_def'   : (ng^2, 2)   deformed GP coords
      'detJ'     : (ng^2,)
      'strain'   : (ng^2, 3)   [exx, eyy, gxy]
      'stress'   : (ng^2, 3)   [sxx, syy, txy]
      'sigma_zz' : (ng^2,)
      'von_mises': (ng^2,)
    """
    mu, nu, planestrain = materialprops
    D = constitutive_matrix(mu, nu, planestrain)

    m1, n1, m2, n2, ngauss, AGtype = eletype
    ngauss = int(ngauss)
    AGtype = int(AGtype)

    edofs = _element_edofs(conn)
    ue = u[edofs]                # (18,)
    ue_nodes = ue.reshape(9, 2)  # (9,2)

    xi_pts, _   = leggauss(ngauss)
    eta_pts, _  = leggauss(ngauss)

    nGP   = ngauss * ngauss
    xi_eta   = np.zeros((nGP, 2))
    xy0      = np.zeros((nGP, 2))
    xy_def   = np.zeros((nGP, 2))
    detJ     = np.zeros(nGP)
    strain   = np.zeros((nGP, 3))
    stress   = np.zeros((nGP, 3))
    sigma_zz = np.zeros(nGP)
    von_mises= np.zeros(nGP)

    k = 0
    for xi in xi_pts:
        for eta in eta_pts:
            # --- Q9 geometry mapping ---
            Ng, dNg_dxi, dNg_deta = q9_geom_shapes_derivs(xi, eta)
            x_gp = 0.0; y_gp = 0.0
            J00=0.0; J01=0.0; J10=0.0; J11=0.0
            for a in range(9):
                xa = coords[conn[a], 0]; ya = coords[conn[a], 1]
                Na = Ng[a]
                x_gp += Na * xa; y_gp += Na * ya
                J00 += dNg_dxi[a]*xa;  J01 += dNg_deta[a]*xa
                J10 += dNg_dxi[a]*ya;  J11 += dNg_deta[a]*ya

            DJ = J00*J11 - J01*J10
            invDJ = 1.0 / DJ
            Jinv00 =  J11*invDJ; Jinv01 = -J01*invDJ
            Jinv10 = -J10*invDJ; Jinv11 =  J00*invDJ

            # --- Displacement field (AG) ---
            if AGtype == 2:
                dN_dxi  = dN_dxi_AGc4(xi, eta, m1, n1, m2, n2)
                dN_deta = dN_deta_AGc4(xi, eta, m1, n1, m2, n2)
                Ndisp   = shape_functions_AGc4(xi, eta, m1, n1, m2, n2)
            else:
                dN_dxi  = dN_dxi_AGc3(xi, eta, m1, n1, m2, n2)
                dN_deta = dN_deta_AGc3(xi, eta, m1, n1, m2, n2)
                Ndisp   = shape_functions_AGc3(xi, eta, m1, n1, m2, n2)

            dNdx = Jinv00 * dN_dxi + Jinv01 * dN_deta
            dNdy = Jinv10 * dN_dxi + Jinv11 * dN_deta

            # --- Strain & stress ---
            B   = build_B(dNdx, dNdy)     # (3x18)
            eps = B @ ue                  # [exx, eyy, gxy]
            sig = D @ eps                 # [sxx, syy, txy]
            sxx, syy, txy = sig[0], sig[1], sig[2]

            # sigma_zz and von Mises
            if planestrain == 1:
                szz = nu * (sxx + syy)
                vm2 = (sxx*sxx + syy*syy + szz*szz
                       - sxx*syy - syy*szz - szz*sxx
                       + 3.0 * txy*txy)
            else:
                szz = 0.0
                vm2 = (sxx*sxx + syy*syy - sxx*syy + 3.0 * txy*txy)

            vm = np.sqrt(vm2) if vm2 > 0.0 else 0.0

            # Deformed GP coords
            u_gp = Ndisp @ ue_nodes   # (2,)
            x_def = x_gp + u_gp[0]; y_def = y_gp + u_gp[1]

            # save
            xi_eta[k, 0] = xi;     xi_eta[k, 1] = eta
            xy0[k, 0]    = x_gp;   xy0[k, 1]    = y_gp
            xy_def[k, 0] = x_def;  xy_def[k, 1] = y_def
            detJ[k]      = DJ
            strain[k,:]  = eps
            stress[k,:]  = sig
            sigma_zz[k]  = szz
            von_mises[k] = vm
            k += 1

    return {
        "xi_eta": xi_eta,
        "xy0": xy0,
        "xy_def": xy_def,
        "detJ": detJ,
        "strain": strain,
        "stress": stress,
        "sigma_zz": sigma_zz,
        "von_mises": von_mises,
    }

def compute_gauss_strain_stress_all_with_deformed(coords, connectivity, etypes, u, materialprops):
    """
    Returns dict with:
      'elem_id','xi_eta','xy0','xy_def','detJ','strain','stress','sigma_zz','von_mises'
    """
    Ne = connectivity.shape[0]
    rows_eid, rows_xi_eta, rows_xy0, rows_xydef = [], [], [], []
    rows_detJ, rows_strain, rows_stress = [], [], []
    rows_szz, rows_vm = [], []

    for e in range(Ne):
        conn = connectivity[e]
        eletype = etypes[e] if not isinstance(etypes, np.ndarray) else etypes[e].tolist()
        res = element_gauss_strain_stress_with_deformed(coords, conn, u, eletype, materialprops)

        nGP = res["strain"].shape[0]
        rows_eid.append(np.full(nGP, e, dtype=int))
        rows_xi_eta.append(res["xi_eta"])
        rows_xy0.append(res["xy0"])
        rows_xydef.append(res["xy_def"])
        rows_detJ.append(res["detJ"])
        rows_strain.append(res["strain"])
        rows_stress.append(res["stress"])
        rows_szz.append(res["sigma_zz"])
        rows_vm.append(res["von_mises"])

    elem_id   = np.concatenate(rows_eid, axis=0)
    xi_eta    = np.concatenate(rows_xi_eta, axis=0)
    xy0       = np.concatenate(rows_xy0, axis=0)
    xy_def    = np.concatenate(rows_xydef, axis=0)
    detJ      = np.concatenate(rows_detJ, axis=0)
    strain    = np.concatenate(rows_strain, axis=0)
    stress    = np.concatenate(rows_stress, axis=0)
    sigma_zz  = np.concatenate(rows_szz, axis=0)
    von_mises = np.concatenate(rows_vm, axis=0)

    return {
        "elem_id": elem_id,
        "xi_eta": xi_eta,
        "xy0": xy0,
        "xy_def": xy_def,
        "detJ": detJ,
        "strain": strain,         # [exx, eyy, gxy]
        "stress": stress,         # [sxx, syy, txy]
        "sigma_zz": sigma_zz,
        "von_mises": von_mises,
    }

